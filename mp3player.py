from tkinter import Tk, Frame, Button, Menu, PhotoImage, Listbox, filedialog, Scrollbar, Label, LabelFrame
import os
import pygame # tkinter doesn't have methods to play music so we will use pygame instead
import openpyxl
import tkinter.messagebox as tmb
import time
from mutagen.mp3 import MP3  # pygame has no methods to check for maximum time in a sound file (except in .wav) 
                # so this library is used. This is not pre-installed in distributions of python
import tkinter.ttk as ttk

# we need a database to store all file names along with their adresses
workbook=openpyxl.load_workbook('Playlist_file.xlsx')
data_sheet=workbook.get_sheet_by_name('Sheet1')
current_row=data_sheet.max_row+1

# initialize the pygame mixed which allows us to play songs
pygame.mixer.init()

# a variable is defined to check which song is currently being used in mixer
current_selection=''

# a variable to hold totall length of a song
song_total_length=0

# a variable to tell how much time has been offset by
offset_time=0

# define functions
def initialise_playlist(): # inputs values of playlist on creation
    for i in range(2,current_row):
        playlist.insert('end', data_sheet.cell(row=i, column=1).value)
    
def add_file(): # Used to add one new file into playlist
    global current_row # global variable is called as changes will be made in this
     
    # can only retreive one file at a time
    fileaddress=filedialog.askopenfilename(initialdir="\\", title="Select a file", 
                                        filetypes=(('MP3 files','*.mp3'),('Ogg files','*.ogg'),
                                                   ('All Files','*.*')))
    # Opens the file dialog box, filetypes dosent work with only one file typeas tuple won't be created
    
    # Extract file name from given file address
    filename=os.path.basename(fileaddress)
    filename=os.path.splitext(filename)[0]
    
    # Add name and corresponding address into data_file
    data_sheet.cell(row=current_row,column=1).value=filename
    data_sheet.cell(row=current_row,column=2).value=fileaddress
    current_row+=1
    
    # insert into playlist
    playlist.insert('end',f'{filename}')
    # buttons cannot be inserted so delete functionality needs to be added by some other method


def add_multiple_files():  # used to add mutiple files into playlist
    global current_row # global variable is called as changes will be made in this
    
    # Gives functionality to Ctrl+Select
    fileaddresses = filedialog.askopenfilenames(initialdir="\\", title="Select required files",
                                             filetypes=(('MP3 files', '*.mp3'), ('Ogg files', '*.ogg'),
                                                        ('All Files', '*.*')))
    # Opens the file dialog box, filetypes dosent work with only one file type as tuple won't be created
    
    for fileaddress in fileaddresses: 
        # Extract file name from given file address
        filename = os.path.basename(fileaddress)
        filename = os.path.splitext(filename)[0]
        
        # Add name and corresponding address into data_file
        data_sheet.cell(row=current_row,column=1).value=filename
        data_sheet.cell(row=current_row,column=2).value=fileaddress
        current_row += 1

        # insert into playlist
        playlist.insert('end', f'{filename}')
        # buttons cannot be inserted so delete functionality needs to be added by some other method
    
def delete_selected(): # Clears currently selected song
    global current_row
    
    for i in range(2,current_row):
        if data_sheet.cell(row=i,column=1).value == playlist.get('anchor'):
            data_sheet.delete_rows(i)
            break 
        # if multiple songs of the same name are present without break all will be deleted, 
        # this makes sure only one instance is deleted
    # Upward shifting is automatically done
    
    # without if first value cannot be deleted
    if current_row==2: # removes a bug caused by range returning no values if start and end are the same here 2
        if data_sheet.cell(row=2, column=1).value == playlist.get('anchor'):
            data_sheet.delete_rows(2)
    
    current_row-=1
            
    playlist.delete('anchor')

def clear_playlist(): # Clears entire playlist 
    global current_row
    
    data_sheet.delete_rows(2,current_row)
    current_row=2
        
    playlist.delete(0,'end')
    
def get_song_address(name):
    for i in range(2,current_row):
        if data_sheet.cell(row=i,column=1).value==name:
            return data_sheet.cell(row=i,column=2).value
        
    if current_row==2:
        return data_sheet.cell(row=2, column=2).value
    
def play_song():
    global current_selection
    global song_total_length
    global offset_time
    
    offset_time=0 # removes changes made by previous songs 
    
    if not playlist.curselection(): # if no selection is made
        playlist.selection_set('active') # select the deafult active one i.e. starting
    
    # remove delete button and place add button
    delete_button.grid_remove()
    add_button.grid(row=0,column=6) 
    
    song_address = get_song_address(playlist.get('active')) # retrieves address from data file
    
    # store currently selected song
    current_selection=playlist.get('active')
    
    # get maximum length of song
    song_mut=MP3(song_address) # returns all information about songfrom location passed
    song_total_length=song_mut.info.length # returns total length of song
    
    # play song
    pygame.mixer.music.load(song_address)
    pygame.mixer.music.play(loops=0)
    playlist.config(selectbackground='green', selectforeground='yellow')
    # loops=0 makes sure song only plays once
    
    # remove play button
    play_button.grid_remove()
    
    # packs status bar
    status_bar.pack(side='bottom', fill='x')
    
    # overwrite play button with the pause button
    pause_button.grid(row=0,column=2)
    
     # inserts song slder frame, done here as now frame will only be inserted once, 
     # not that it would return an error if done in play ime
    song_slider_frame.grid(row=2,column=0)
    song_slider.config(to=song_total_length)
    proper_total_time=time.strftime('%H:%M:%S', time.gmtime(song_total_length))
    total_length_label.config(text=proper_total_time)
    
    # Time for status bar
    play_time()
    
    
def pause_song():
    # pauses current song
    pygame.mixer.music.pause()
    playlist.config(selectbackground='white', selectforeground='red')
    
    # overwrite pause button with play button if selection changed otherwise to unpause button
    pause_button.grid_remove()
    
    if(current_selection != playlist.get('active')):
        play_button.grid(row=0,column=2)
    else:
        unpause_button.grid(row=0,column=2)
    
def unpause_song():
    if(current_selection != playlist.get('active')):
        play_song()
    else:
        pygame.mixer.music.unpause()
        playlist.config(selectbackground='green', selectforeground='yellow')
        unpause_button.grid_remove()
        pause_button.grid(row=0,column=2)
    
def forward_song():
    global offset_time
    
    # skips forward by 10 seconds
    current_time=offset_time+(pygame.mixer.music.get_pos()/1000)
    
    if(current_time+10>=song_total_length):
        pass
    else:
        pygame.mixer.music.set_pos(current_time+10)
        offset_time += 10

def rewind_song():
    # rewinds by 10 seconds
    global offset_time

    # goes backwards by 10 seconds
    current_time = offset_time+(pygame.mixer.music.get_pos()/1000)

    if(current_time-10 <= 0):
        pass
    else:
        pygame.mixer.music.set_pos(current_time-10)
        offset_time -= 10

def next_song():
    # goes to next song
    playlist_size=playlist.size()
    
    if playlist.curselection(): # only proceed if current selection tuple is not empty
        current_index=playlist.curselection()[0]
        if current_index+1<playlist_size: #keep going forward unless you are at your last song
            playlist.selection_clear(0,'end') 
            # without clearing multiple selections would be made so first we delete previous selection 
            # before updating with our own
            playlist.selection_set((current_index+1,)) # without this the selection bar wouldn't appear
            playlist.activate(current_index+1) # switches anchor to next index
            play_song()
        else:
            tmb.showerror('Error Message',message='No more songs are available')
    else:
        tmb.showerror(message='No selection was made')

def previous_song():
    # goes to previous song
    if playlist.curselection():  # only proceed if current selection tuple is not empty
        current_index=playlist.curselection()[0]
        if current_index>0: # keep going back unless you are at the first song
            playlist.selection_clear(0,'end') # current selections are removed
            playlist.selection_set((current_index-1,))
            playlist.activate(current_index-1)
            play_song()
        else:
            tmb.showerror('Error Message', message='No more songs are available')
    else:
        tmb.showerror(message='No selection was made')

def stop_playing(): 
    global offset_time
    
    # stops mixer from playing further song and returns to default screen and activates delete button
    pygame.mixer.music.stop()
    playlist.config(selectbackground='yellow', selectforeground='green')
    
    # grid_remove doesn't return an error if a widget isn't oresent in the file
    pause_button.grid_remove()
    unpause_button.grid_remove()
    play_button.grid(row=0,column=2)
     
    add_button.grid_remove()
    delete_button.grid(row=0,column=6)
    
    # handle status bar after stop
    offset_time=0
    status_bar.config(text='')
    
    # remove silder
    song_slider_frame.grid_forget()
    
    # removes status bar
    status_bar.pack_forget()
  
def keep_checking_for_changes(last_selected):
    if last_selected != playlist.get('active'):
        replay_button.grid_remove()
        play_button.grid(row=0,column=2)
    
    status_bar.after(1000,lambda: keep_checking_for_changes(last_selected)) 
    # there is no need for after() to be associated with the label "status_bar", agin is a tkinter function and
    # works with any instance of tkinter
    
def play_time():
    global currently_playing 
    time_of_song=offset_time+(pygame.mixer.music.get_pos()/1000) # returns time of song in milliseconds
    
    # To make this more informativr, we will convert data in milliseconds to proper time format
    proper_time=time.strftime('%H:%M:%S', time.gmtime(time_of_song))
    proper_total_time=time.strftime('%H:%M:%S', time.gmtime(song_total_length))
    # division by 1000 was done because gmtime() expects data in seconds and not milliseconds
    
    # Put this in status bar to show the user
    if time_of_song>=1:
        status_bar.config(text=f"{playlist.get('active')}: {proper_time}/{proper_total_time} ")
        
    #To show information on song slider
    current_length_label.config(text=proper_time)
    song_slider.config(value=time_of_song)
        
    # if song finishes show a replay button that is replaced again by play button when choice changes
    if int(time_of_song) == int(song_total_length):
        play_button.grid_remove()
        song_slider_frame.grid_remove()
        replay_button.grid(row=0,column=2)
        status_bar.pack_forget()
        playlist.config(selectbackground='yellow', selectforeground='green')
        keep_checking_for_changes(playlist.get('active'))
        currently_playing=False
        
    status_bar.after(1000,play_time) # this will call the function again every 1000 millisecons -> 1 second

def replay_song(): # replays song
    play_song()
    
    # remove replay button and place play button
    replay_button.grid_remove()
    playlist.config(selectbackground='green', selectforeground='yellow')
    play_button.grid(row=0,column=2)
    
def volume_adjustment(slider_position_in_string):
    # this function is executed every time the slider is manipulated
    # and is passed the current value of the slider's position
    # but slider_position is a string value, so instead of converting it which may cause loss of value 
    # we instead get the position as a floating point literal using .get()
    pygame.mixer.music.set_volume(volume_slider.get()) 
    
    # update label
    volume_label.config(text=round(volume_slider.get()*100,2))
    
def slider_adjustment(slider_position_in_string):
    global offset_time
    # edited the code of play_song to create a new instance of song selected but set offset to 
    # current position of the slider so user thinks this is the same song whreas in actuality it is 
    # another instance of the same song staring from end of current instance to end
    
    offset_time = song_slider.get()  # removes changes made by previous songs
    
    song_address = get_song_address(playlist.get('active')) # retrieves address from data file

    # store currently selected song
    current_selection=playlist.get('active')
    
    # total length is independant of song length hence was no altered
    
    # play song
    pygame.mixer.music.load(song_address)
    pygame.mixer.music.play(loops=0, start=song_slider.get())
    playlist.config(selectbackground='green', selectforeground='yellow')
    
    # Time for status bar
    play_time()

root=Tk()
root.geometry(f'{root.winfo_screenwidth()}x{root.winfo_screenheight()}')
root.title('MP3 Player')
root.wm_iconbitmap('Iconshock-Orchestra-Harp.ico')

# Creating the menu bar
menu_bar= Menu(root)

# create two submenues that will insert files one by one or multiple in one go
insertion_menu = Menu(menu_bar, tearoff=False)
insertion_menu.add_command(label='Insert Single File', command=add_file)
insertion_menu.add_command(label='Insert Multiple Files', command=add_multiple_files)

# Creating an add file option which will contain a single insert and multiple insert options
menu_bar.add_cascade(label="Add New", menu=insertion_menu)

# Creating a deletion menu
deletion_menu=Menu(menu_bar, tearoff=False)
deletion_menu.add_command(label='Delete Selected Song', command=delete_selected)
deletion_menu.add_command(label='Clear Playlist', command=clear_playlist)

# adding deletion menu to menu bar
menu_bar.add_cascade(label='Delete File', menu=deletion_menu)

# creating frames
main_frame=Frame(root)
controller_frame=Frame(main_frame)
playlist_frame=Frame(main_frame)
volume_frame=LabelFrame(playlist_frame, text='Volume') # labelframe is a frame with a border made to resemble a label
song_slider_frame=Frame(main_frame)

# creating a listbox to show song names
playlist=Listbox(playlist_frame, bg='black', fg='white', width=100, height=20, 
                 relief='sunken', borderwidth=10, selectbackground='yellow', selectforeground='green')
playlist.pack(pady=10, side='left')
initialise_playlist()

#adding scrollbar to playlist
scroll=Scrollbar(playlist_frame)
playlist.config(yscrollcommand=scroll.set)
scroll.config(command=playlist.yview)
scroll.pack(fill='y', side='left')

# creating and adding controls
play_button_icon = PhotoImage(file="play-icon.png")
play_button=Button(controller_frame, image=play_button_icon, borderwidth=0, command=play_song, activebackground='green')
play_button.grid(row=0,column=2)
# Will appear by default, it will overwrite pause or unpause once selection is altered

unpause_button_icon = PhotoImage(file="play-icon.png")
unpause_button = Button(controller_frame, image=play_button_icon, borderwidth=0, command=unpause_song, activebackground='green')
# this will have same icon as play button and will give the user the illusion that the pause button 
# is being replaced by the play button wheras it is actually being changed into the unpause button

pause_button_icon = PhotoImage(file="pause-icon.png")
pause_button = Button(controller_frame, image=pause_button_icon,
                      borderwidth=0, command=pause_song, activebackground='green')
# Pause will override play when video is playing

forward_button_icon = PhotoImage(file="forward-icon.png")
forward_button = Button(controller_frame, image=forward_button_icon,
                        borderwidth=0, command=forward_song, activebackground='green')
forward_button.grid(row=0,column=4)

rewind_button_icon = PhotoImage(file="rewind-icon.png")
rewind_button = Button(controller_frame, image=rewind_button_icon,
                       borderwidth=0, command=rewind_song, activebackground='green')
rewind_button.grid(row=0,column=1)

replay_button_icon = PhotoImage(file="replay-icon.png")
replay_button=Button(controller_frame, image=replay_button_icon, borderwidth=0, command= replay_song, activebackground='green')
# This will override once music is done playing

previous_button_icon = PhotoImage(file="first-track-icon.png")
previous_button = Button(controller_frame, image=previous_button_icon,
                         borderwidth=0, command=previous_song, activebackground='green')
previous_button.grid(row=0,column=0)

next_button_icon = PhotoImage(file="last-track-icon.png")
next_button=Button(controller_frame, image=next_button_icon, borderwidth=0, command=next_song, activebackground='green')
next_button.grid(row=0,column=5)

add_button_icon = PhotoImage(file="add-icon.png")
add_button=Button(controller_frame, image=add_button_icon, borderwidth=0, command=add_multiple_files, activebackground='green')
add_button.grid(row=0,column=6)

stop_button_icon = PhotoImage(file="stop-icon.png")
stop_button=Button(controller_frame, image=stop_button_icon, borderwidth=0, command=stop_playing, activebackground='green')
stop_button.grid(row=0,column=3)

delete_button_icon = PhotoImage(file="delete-icon.png")
delete_button=Button(controller_frame, image=delete_button_icon, borderwidth=0, command=delete_selected)
# delete_button.grid(row=0,column=6)
# this will override the add new button when a song is sewlected,giving the user theoption to either play 
# or delete it, it gets overriden once play is pressed

# create a status bar
status_bar=Label(root, bg='#941903', foreground='white', relief='sunken', borderwidth=5)
# status bar will remain hidden until a song starts playing and 
# once again will be hidden once song is done playing

# volume adjustment bar
# pygame volume control goes from 0 to 1, so we need to adjust for that
volume_slider=ttk.Scale(volume_frame, orient='vertical', from_=1, to=0, command=volume_adjustment, value=0.5, length=310)
# a ttk scale is different from a normal tkinter scale as it allows selection of intermediate values 
# as opposed to only integral values as was the case with tkinter scales
# value defines initial position of slider
volume_slider.grid(row=1,column=0)
volume_frame.pack(side='left', fill='y')

# volume indicator
volume_label=Label(volume_frame, text=volume_slider.get()*100)
volume_label.grid(row=0, column=0)

# Making a song slider
song_slider=ttk.Scale(song_slider_frame, orient='horizontal', length=840, command=slider_adjustment)
song_slider.grid(pady=20,row=0,column=1)

# position labels for song slider
total_length_label=Label(song_slider_frame)
current_length_label=Label(song_slider_frame)

#packing labels into frame
total_length_label.grid(row=0,column=2)
current_length_label.grid(row=0,column=0)

# packing subframes into main frame
playlist_frame.grid(row=0,column=0)
controller_frame.grid(row=1,column=0)
root.config(menu=menu_bar)  # configuring menu bar into the root window

root.config(bg='yellow')

# packing frames and executing loops
main_frame.pack()
root.mainloop()
workbook.save('Playlist_file.xlsx')
